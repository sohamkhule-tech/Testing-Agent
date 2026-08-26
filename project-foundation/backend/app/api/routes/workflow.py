"""
Workflow API Routes

Exposes per-stage workflow data that was previously only available
as filesystem artifacts.  Each endpoint reads the relevant JSON file
from the run's workspace directory and returns it via the existing
Pydantic schemas.

All existing ``/api/v1/runs/*`` endpoints remain fully backward compatible.
"""

import io
import re
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse, StreamingResponse

from app.dependencies import get_trigger_service
from app.exceptions import NotFoundError
from app.logging import get_logger
from app.services import TriggerService
from app.utils import load_file

logger = get_logger("api.workflow")

router = APIRouter(prefix="/runs", tags=["Workflow"])


# ===================================================================
# Helpers
# ===================================================================


async def _get_run_workspace(
    run_id: UUID,
    service: TriggerService,
) -> Path:
    """Return the ``Path`` to the run's workspace root.

    Raises ``HTTPException(404)`` if the run does not exist.
    """
    try:
        entity = await service.get_run(run_id)
    except NotFoundError:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Run not found: {run_id}",
        )
    return Path(entity.workspace_path)


def _load_json(path: Path) -> dict | list | None:
    """Safely load a JSON file, returning ``None`` if missing."""
    if not path.exists():
        return None
    try:
        from app.utils.json_utils import loads
        return loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("failed_to_load %s: %s", path, exc)
        return None


# ===================================================================
# Phase 2 — Workflow Status
# ===================================================================


@router.get(
    "/{run_id}/workflow",
    summary="Get complete workflow status",
    description="Returns the overall workflow state including all stage statuses, timing, and errors.",
    responses={
        200: {"description": "Workflow status"},
        404: {"description": "Run not found"},
    },
)
async def get_workflow(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the full workflow execution state for a run.

    Combines run metadata with per-stage file-existence checks to
    determine which stages have completed, which are pending, and
    which have failed.
    """
    workspace = await _get_run_workspace(run_id, service)

    # Base run metadata
    metadata = await service.get_metadata(run_id)

    # Map stage names to their contract file paths and agent labels
    # NOTE: use specific files (not directories) to avoid false positives when dir exists but stage failed
    stages_config = [
        ("trigger",       workspace / "contracts" / "test-run-request.json",       "Trigger Agent"),
        ("crawler",       workspace / "contracts" / "crawl-package.json",          "Crawler Agent"),
        ("inventory",     workspace / "contracts" / "inventory.json",              "Inventory Aggregator"),
        ("test_design",   workspace / "contracts" / "test-plan.json",              "Test Design Agent"),
        ("human_review",  workspace / "contracts" / "approved-test-plan.json",     "Human Review"),
        ("code_generation", workspace / "artifacts" / "generated-tests" / "playwright" / "code-generation-metadata.json", "Code Generation Agent"),
        ("execution",     workspace / "artifacts" / "generated-tests" / "playwright" / "test-results" / "results.json", "Execution Agent"),
    ]

    stages = []
    completed_count = 0
    errors: list[str] = []

    is_workflow_completed = (
        (hasattr(metadata.status, "value") and metadata.status.value == "completed")
        or metadata.status == "completed"
        or metadata.current_stage == "completed"
    )

    # Handle "resuming_from_X" current_stage — treat it as stage X being 'running'
    raw_current_stage = metadata.current_stage or ""
    is_resuming = raw_current_stage.startswith("resuming_from_")
    # The effective stage that is actively running (during resume or normal)
    effective_running_stage = (
        raw_current_stage[len("resuming_from_"):] if is_resuming else raw_current_stage
    )

    # Determine overall failure status
    is_failed = (
        (hasattr(metadata.status, "value") and metadata.status.value == "failed")
        or metadata.status == "failed"
    )

    for stage_name, file_path, agent_label in stages_config:
        exists = file_path.exists()
        is_current = effective_running_stage == stage_name if stage_name != "trigger" else False

        if stage_name == "trigger" or is_workflow_completed:
            # Trigger stage always completes if we have a run entity, and if workflow is completed, all stages are completed
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title() if stage_name != "trigger" else "Project Setup",
                "agent": agent_label,
                "status": "completed",
                "has_data": exists or is_workflow_completed,
            })
            completed_count += 1
        elif exists:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "completed",
                "has_data": True,
            })
            completed_count += 1
        elif is_failed and is_current:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "failed",
                "has_data": False,
            })
            errors.append(f"{stage_name}: {metadata.error or 'Unknown error'}")
        elif is_current:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "running",
                "has_data": False,
            })
        else:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "pending",
                "has_data": False,
            })

    total_stages = len(stages_config)

    return {
        "run_id": str(run_id),
        "overall_status": metadata.status.value if hasattr(metadata.status, "value") else metadata.status,
        "current_stage": metadata.current_stage,
        "progress_percent": metadata.progress_percent,
        "started_at": metadata.created_at.isoformat() if hasattr(metadata.created_at, "isoformat") else str(metadata.created_at),
        "updated_at": metadata.updated_at.isoformat() if hasattr(metadata.updated_at, "isoformat") else str(metadata.updated_at),
        "total_stages": total_stages,
        "completed_stages": completed_count,
        "pending_stages": total_stages - completed_count,
        "stages": stages,
        "errors": errors if errors else None,
        "message": metadata.message,
    }


# ===================================================================
# Phase 3 — Crawler
# ===================================================================


@router.get(
    "/{run_id}/crawler",
    summary="Get crawler results",
    description="Returns the complete crawl-package.json output for a run.",
    responses={
        200: {"description": "Crawler results"},
        404: {"description": "Run or crawler data not found"},
    },
)
async def get_crawler_results(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the ``CrawlPackage`` for a completed crawl.

    Reads ``contracts/crawl-package.json`` from the run workspace
    and validates it against the ``CrawlPackage`` schema.
    """
    workspace = await _get_run_workspace(run_id, service)
    crawl_path = workspace / "contracts" / "crawl-package.json"

    data = _load_json(crawl_path)
    if data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Crawler data not found for run: {run_id}. The crawl stage may not have completed.",
        )

    # Validate and return using the existing schema
    from app.schemas.crawler import CrawlPackage
    crawl_package = CrawlPackage(**data)

    return crawl_package.model_dump(mode="json")


@router.get(
    "/{run_id}/screenshots-list",
    summary="List screenshots captured for a run",
    description="Returns a list of all screenshots captured during crawling/execution for a run.",
)
async def list_screenshots(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    shots_dir = workspace / "screenshots"
    if not shots_dir.exists():
        shots_dir = workspace / "artifacts" / "screenshots"

    screenshots = []
    if shots_dir.exists():
        for i, img_path in enumerate(sorted(shots_dir.glob("*.png"), key=lambda p: p.stat().st_mtime)):
            screenshots.append({
                "id": f"ss-{i}",
                "filename": img_path.name,
                "url": f"/api/v1/runs/{run_id}/screenshots/{img_path.name}",
                "title": img_path.name,
                "timestamp": datetime.fromtimestamp(img_path.stat().st_mtime, tz=timezone.utc).isoformat(),
            })

    return {
        "run_id": str(run_id),
        "screenshots": screenshots,
        "count": len(screenshots),
    }


# ===================================================================
# Phase 4 — Inventory
# ===================================================================


@router.get(
    "/{run_id}/inventory",
    summary="Get inventory results",
    description="Returns the aggregated inventory for a run.",
    responses={
        200: {"description": "Inventory data"},
        404: {"description": "Run or inventory data not found"},
    },
)
async def get_inventory(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the ``Inventory`` for a completed inventory aggregation.

    Reads ``contracts/inventory.json`` from the run workspace.
    """
    workspace = await _get_run_workspace(run_id, service)
    inv_path = workspace / "contracts" / "inventory.json"

    data = _load_json(inv_path)
    if data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Inventory data not found for run: {run_id}.",
        )

    from app.schemas.inventory import Inventory
    inventory = Inventory(**data)
    return inventory.model_dump(mode="json")


# ===================================================================
# Phase 5 — Test Plan
# ===================================================================


@router.get(
    "/{run_id}/test-plan",
    summary="Get test plan results",
    description="Returns the AI-generated test plan for a run.",
    responses={
        200: {"description": "Test plan data"},
        404: {"description": "Run or test plan not found"},
    },
)
async def get_test_plan(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the ``TestPlan`` generated by the Test Design Agent.

    Reads ``contracts/test-plan.json`` from the run workspace.
    """
    workspace = await _get_run_workspace(run_id, service)
    plan_path = workspace / "contracts" / "test-plan.json"

    data = _load_json(plan_path)
    if data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Test plan not found for run: {run_id}.",
        )

    from app.schemas.test_plan import TestPlan
    test_plan = TestPlan(**data)
    return test_plan.model_dump(mode="json")


from fastapi.responses import FileResponse, StreamingResponse
import io


@router.get(
    "/{run_id}/test-plan/export",
    summary="Export test plan as Excel (.xlsx)",
    description="Downloads the test plan as a formatted Excel file with scenarios, modules, and metadata.",
)
async def export_test_plan_xlsx(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
):
    workspace = await _get_run_workspace(run_id, service)
    plan_path = workspace / "contracts" / "test-plan.json"

    data = _load_json(plan_path)
    if data is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test plan not found")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    app_sum = data.get("application_summary", {}) or {}
    ws_summary.append(["Field", "Value"])
    ws_summary.append(["Application Name", app_sum.get("name", "")])
    ws_summary.append(["Total Pages", app_sum.get("total_pages", 0)])
    ws_summary.append(["Total Forms", app_sum.get("total_forms", 0)])
    ws_summary.append(["Total APIs", app_sum.get("total_apis", app_sum.get("totalApis", 0))])
    ws_summary.append(["Auth Required", app_sum.get("authentication_required", False)])
    ws_summary.append(["Auth Method", app_sum.get("auth_method", "none")])
    ws_summary.append(["Total Scenarios", data.get("coverage_summary", {}).get("total_scenarios", 0) if isinstance(data.get("coverage_summary"), dict) else 0])
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    # ── Sheet 2: Scenarios ──
    ws_scenarios = wb.create_sheet("Scenarios")
    headers = ["ID", "Title", "Description", "Module", "Priority", "Category", "Risk Level", "Target Page", "Test Steps", "Expected Result", "Preconditions", "Required Data", "Tags", "Dependencies"]
    ws_scenarios.append(headers)
    for cell in ws_scenarios[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    modules = data.get("modules", []) or []
    seen_ids = set()
    for mod in modules:
        scenarios = mod.get("scenarios", []) or []
        for sc in scenarios:
            meta = sc.get("metadata", {}) or {}
            sid = meta.get("id", "") or ""
            if sid in seen_ids:
                continue
            seen_ids.add(sid)
            ws_scenarios.append([
                sid,
                meta.get("title", ""),
                meta.get("description", ""),
                meta.get("module", mod.get("name", "")),
                meta.get("priority", "medium"),
                meta.get("category", "functional"),
                meta.get("risk_level", "medium"),
                meta.get("target_page", ""),
                "\n".join(meta.get("test_steps", []) or []),
                meta.get("expected_result", ""),
                "\n".join(meta.get("preconditions", []) or []),
                ", ".join(meta.get("required_test_data", []) or []),
                ", ".join(meta.get("tags", []) or []),
                ", ".join(meta.get("dependencies", []) or []),
            ])

    for col in ws_scenarios.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_scenarios.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 3: Modules ──
    ws_modules = wb.create_sheet("Modules")
    ws_modules.append(["Module Name", "Description", "Scenario Count", "Pages"])
    for cell in ws_modules[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for mod in modules:
        ws_modules.append([
            mod.get("name", ""),
            mod.get("description", ""),
            len(mod.get("scenarios", []) or []),
            ", ".join(mod.get("pages", []) or []),
        ])
    ws_modules.column_dimensions["A"].width = 25
    ws_modules.column_dimensions["B"].width = 40
    ws_modules.column_dimensions["C"].width = 15
    ws_modules.column_dimensions["D"].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=test-plan-{run_id}.xlsx"},
    )


# ===================================================================
# Phase 6 — Human Review
# ===================================================================


@router.get(
    "/{run_id}/review",
    summary="Get human review results",
    description="Returns the human review metadata and approved test plan for a run.",
    responses={
        200: {"description": "Review data"},
        404: {"description": "Run or review data not found"},
    },
)
async def get_human_review(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return review metadata and the approved test plan.

    Reads ``contracts/review-metadata.json`` and optionally
    ``contracts/approved-test-plan.json`` from the run workspace.
    """
    workspace = await _get_run_workspace(run_id, service)

    # Load review metadata
    review_meta_path = workspace / "contracts" / "review-metadata.json"
    review_data = _load_json(review_meta_path)

    if review_data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Review data not found for run: {run_id}.",
        )

    # Load approved plan if it exists
    approved_plan_path = workspace / "contracts" / "approved-test-plan.json"
    approved_plan = _load_json(approved_plan_path)

    result = {
        "run_id": str(run_id),
        "review_metadata": review_data,
    }
    if approved_plan:
        result["approved_test_plan"] = approved_plan

    return result


# ===================================================================
# Phase 7 — Execution Results
# ===================================================================


def _parse_test_results_from_folders(test_results_dir: Path, pw_dir: Path | None = None) -> list[dict]:
    """Parse individual test results from Playwright's test-results directory."""
    tests = []

    last_run_path = test_results_dir / ".last-run.json"
    failed_ids: set[str] = set()
    if last_run_path.exists():
        try:
            lr = _load_json(last_run_path)
            if lr and isinstance(lr, dict):
                failed_ids = set(lr.get("failedTests", []))
        except Exception:
            pass

    # Try to find results.json (Playwright JSON reporter output)
    results_json = test_results_dir / "results.json"
    if results_json.exists():
        try:
            data = _load_json(results_json)
            if data and isinstance(data, dict):
                status_map = {
                    "passed": "passed", "failed": "failed", "timedOut": "failed",
                    "skipped": "skipped", "interrupted": "failed",
                }

                def _walk_suites(suites: list) -> None:
                    # Playwright JSON nests specs under describe-blocks inside file suites.
                    # Top-level file suite has specs=[] and suites=[describe_block].
                    # Must recurse into suites to find actual specs.
                    for suite in suites or []:
                        for spec in suite.get("specs", []):
                            for test in spec.get("tests", []):
                                results_list = test.get("results", [])
                                status = "skipped"
                                duration_ms = 0
                                error_msg = None
                                if results_list:
                                    last = results_list[-1]
                                    status = status_map.get(last.get("status", "skipped"), "skipped")
                                    duration_ms = sum(r.get("duration", 0) for r in results_list)
                                    if status == "failed":
                                        err = last.get("error", {})
                                        error_msg = err.get("message") if isinstance(err, dict) else None
                                tests.append({
                                    "id": f"{suite.get('file', '')}-{spec.get('title', '')}-{len(tests)}",
                                    "name": spec.get("title", "Unknown"),
                                    "file": spec.get("file") or suite.get("file", ""),
                                    "status": status,
                                    "duration": duration_ms,
                                    "error": error_msg,
                                    "browser": test.get("projectName"),
                                    "timestamp": "",
                                })
                        _walk_suites(suite.get("suites", []))

                _walk_suites(data.get("suites", []))
            if tests:
                return tests
        except Exception:
            pass

    # Fallback 2: parse test-results subfolders.
    # Use os.scandir so DirEntry.is_dir() reads the cached WIN32_FIND_DATA
    # attribute — this works even when the folder path exceeds MAX_PATH (260)
    # because it avoids the GetFileAttributesW syscall that pathlib.Path.is_dir()
    # would make on the long path.
    import os as _os
    if test_results_dir.exists():
        try:
            entries = sorted(_os.scandir(str(test_results_dir)), key=lambda e: e.name)
        except OSError:
            entries = []
        for entry in entries:
            if entry.name.startswith(".") or not entry.is_dir():
                continue
            parts = entry.name.split("-")
            name_parts = parts[2:-2] if len(parts) > 4 else parts[1:]
            name = " ".join(name_parts).replace("-", " ").strip() or entry.name

            status = "failed"
            error_msg = None

            stderr_file = test_results_dir / entry.name / "error.txt"
            if stderr_file.exists():
                try:
                    error_msg = stderr_file.read_text(encoding="utf-8", errors="replace")[:200]
                except Exception:
                    pass

            tests.append({
                "id": entry.name,
                "name": name,
                "file": "",
                "status": status,
                "duration": None,
                "error": error_msg,
                "browser": parts[-1] if parts else "chromium",
                "timestamp": "",
            })

    if tests:
        return tests

    # Fallback 3: Parse generated spec files directly if test-results was empty/missing
    # CRITICAL: These tests were NEVER EXECUTED - mark them as "not_executed", NEVER "passed"
    if pw_dir:
        tests_dir = pw_dir / "tests"
        if tests_dir.exists():
            import re
            for spec in sorted(tests_dir.glob("*.spec.ts")):
                try:
                    content = spec.read_text(encoding="utf-8", errors="replace")
                    for idx, match in enumerate(re.finditer(r"test\(\s*['\"]([^'\"]+)['\"]", content)):
                        title = match.group(1)
                        tests.append({
                            "id": f"{spec.name}-{title}-{idx}",
                            "name": title,
                            "file": f"tests/{spec.name}",
                            "status": "not_executed",
                            "duration": None,
                            "error": "Test was generated but Playwright produced no execution results. The test was never actually run.",
                            "browser": "chromium",
                            "timestamp": "",
                        })
                except Exception:
                    pass

    return tests


@router.get(
    "/{run_id}/execution",
    summary="Get test execution results",
    description="Returns test execution results, metrics, and status for a run.",
)
async def get_execution_results(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)

    exec_artifacts = workspace / "artifacts" / "generated-tests" / "execution-artifacts"
    pw_dir = workspace / "artifacts" / "generated-tests" / "playwright"
    test_results_dir = pw_dir / "test-results"

    # Load summary
    summary_path = exec_artifacts / "reports" / "execution-summary.json"
    summary = _load_json(summary_path) if summary_path.exists() else None

    # Load metrics
    metrics_path = exec_artifacts / "execution-metrics.json"
    metrics = _load_json(metrics_path) if metrics_path.exists() else None

    # Load metadata
    meta_path = exec_artifacts / "execution-metadata.json"
    exec_meta = _load_json(meta_path) if meta_path.exists() else None

    # Parse test results from folders / results.json / spec files
    tests = _parse_test_results_from_folders(test_results_dir, pw_dir)

    # Also try execution-artifacts/artifact-index.json
    artifact_index_path = exec_artifacts / "artifact-index.json"
    artifact_index = _load_json(artifact_index_path) if artifact_index_path.exists() else None
    if artifact_index and isinstance(artifact_index, dict):
        index_tests = artifact_index.get("tests", [])
        if index_tests and not tests:
            tests = index_tests

    # Build summary stats
    total = len(tests)
    passed = sum(1 for t in tests if t.get("status") == "passed")
    failed = sum(1 for t in tests if t.get("status") == "failed")
    skipped = sum(1 for t in tests if t.get("status") == "skipped")
    not_executed = sum(1 for t in tests if t.get("status") == "not_executed")
    
    # CRITICAL: Pass rate should only count EXECUTED tests (passed + failed + skipped)
    # NOT tests that were never run (not_executed)
    executed_tests = passed + failed + skipped
    pass_rate = (passed / executed_tests * 100) if executed_tests > 0 else 0.0

    # If metrics has actual data and folder parsing is empty, prefer metrics
    if metrics and isinstance(metrics, dict) and metrics.get("total_tests", 0) > 0 and total == 0:
        total = metrics.get("total_tests", 0)
        passed = metrics.get("tests_passed", 0)
        failed = metrics.get("tests_failed", 0)
        skipped = metrics.get("tests_skipped", 0)
        pass_rate = metrics.get("pass_rate", 0.0)

    execution_complete = exec_meta is not None or summary is not None or total > 0
    
    # Determine execution status and classification
    return_code = (exec_meta or {}).get("return_code")
    classification = (exec_meta or {}).get("classification")

    # Normalize legacy classification values written before the taxonomy was
    # clarified (playwright_timeout → execution_timeout).
    if classification == "playwright_timeout":
        classification = "execution_timeout"

    execution_status = "completed"
    executed_count = sum(1 for t in tests if t.get("status") in ("passed", "failed", "skipped"))
    if classification == "execution_timeout":
        # A wall-clock kill that still produced complete, usable results is a
        # completed execution with failures — NOT a timeout. This heals runs
        # that were previously tagged as timeouts despite full results.
        if executed_count > 0 or total > 0:
            classification = "test_execution_completed_with_failures"
            execution_status = "completed"
        else:
            execution_status = "execution_timeout"
    elif classification == "infrastructure_failure":
        execution_status = "infrastructure_failure"
    elif classification == "test_execution_completed_with_failures":
        execution_status = "completed"
    elif return_code is not None and return_code < 0 and executed_count == 0:
        execution_status = "infrastructure_failure"
    elif summary:
        execution_status = summary.get("status", "completed")
    elif execution_complete:
        execution_status = "completed"
    else:
        execution_status = "pending"

    return {
        "run_id": str(run_id),
        "status": execution_status,
        "classification": classification,
        "execution_complete": execution_complete,
        "duration_seconds": (summary or {}).get("duration_seconds") or (exec_meta or {}).get("duration_seconds"),
        "return_code": return_code,
        "tests": tests,
        "summary": {
            "total": total,
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
            "not_executed": not_executed,
            "executed": executed_tests,
            "pass_rate": round(pass_rate, 1) if executed_tests > 0 else None,
        },
        "metrics": metrics,
        "execution_metadata": exec_meta,
    }


@router.get(
    "/{run_id}/reports",
    summary="Get execution reports",
    description="Returns full execution reports including failure analysis, metrics, and summary.",
)
async def get_execution_reports(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    exec_artifacts = workspace / "artifacts" / "generated-tests" / "execution-artifacts"

    reports_dir = exec_artifacts / "reports"
    summary = _load_json(reports_dir / "execution-summary.json") if (reports_dir / "execution-summary.json").exists() else None
    failure_report = _load_json(reports_dir / "failure-report.json") if (reports_dir / "failure-report.json").exists() else None
    metrics_report = _load_json(reports_dir / "metrics-report.json") if (reports_dir / "metrics-report.json").exists() else None
    failure_analysis_path = exec_artifacts / "failure-analysis" / "failure-analysis.json"
    failure_analysis = _load_json(failure_analysis_path) if failure_analysis_path.exists() else None

    has_html_report = (exec_artifacts / "reports" / "playwright-report").exists() or \
                      (workspace / "artifacts" / "generated-tests" / "playwright" / "playwright-report").exists()

    return {
        "run_id": str(run_id),
        "has_data": any(x is not None for x in [summary, failure_report, metrics_report]),
        "execution_summary": summary,
        "failure_report": failure_report,
        "metrics_report": metrics_report,
        "failure_analysis": failure_analysis,
        "has_html_report": has_html_report,
        "has_allure_report": _get_allure_report_dir(workspace).joinpath("index.html").exists(),
    }


# ===================================================================
# Allure Report
# ===================================================================


def _get_allure_report_dir(workspace: Path) -> Path:
    """Return the canonical per-run Allure report directory."""
    exec_artifacts = workspace / "artifacts" / "generated-tests" / "execution-artifacts"
    return exec_artifacts / "reports" / "allure-report"


def _resolve_report_file(report_dir: Path, relative_path: str) -> Path | None:
    """Resolve a path inside the Allure report dir, guarding against traversal."""
    report_root = report_dir.resolve()
    target = (report_root / relative_path).resolve()
    try:
        target.relative_to(report_root)
    except ValueError:
        return None
    return target


def _existing_file_path(path: Path) -> Path | None:
    """Return a usable file path, including Windows long-path support."""
    if path.is_file():
        return path

    if path.drive:
        resolved = path.resolve()
        raw = str(resolved)
        if not raw.startswith("\\\\?\\"):
            if raw.startswith("\\\\"):
                raw = "\\\\?\\UNC\\" + raw.lstrip("\\")
            else:
                raw = "\\\\?\\" + raw
        long_path = Path(raw)
        if long_path.is_file():
            return long_path

    return None


@router.get(
    "/{run_id}/report/status",
    summary="Get Allure report status",
    description="Returns whether an Allure report is available for a run and where it lives.",
    responses={
        200: {"description": "Allure report status"},
        404: {"description": "Run not found"},
    },
)
async def get_allure_report_status(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    report_dir = _get_allure_report_dir(workspace)

    index_exists = report_dir.joinpath("index.html").exists()

    if not index_exists:
        # Check if test execution results exist and attempt auto-generation
        pw_dir = workspace / "artifacts" / "generated-tests" / "playwright"
        test_results_dir = pw_dir / "test-results"
        fallback_tests = _parse_test_results_from_folders(test_results_dir, pw_dir)
        if fallback_tests:
            from app.execution.allure_report_generator import AllureReportGenerator
            allure_results_dir = pw_dir / "allure-results"
            generator = AllureReportGenerator()
            generator.generate(
                results_dir=allure_results_dir,
                output_path=report_dir,
                project_path=pw_dir if pw_dir.exists() else None,
                fallback_test_results=fallback_tests,
                force_rebuild=True,
            )
            index_exists = report_dir.joinpath("index.html").exists()

    if index_exists:
        status_value = "generated"
    elif report_dir.exists():
        status_value = "failed"
    else:
        status_value = "unavailable"

    return {
        "run_id": str(run_id),
        "status": status_value,
        "report_available": index_exists,
        "report_path": str(report_dir) if report_dir.exists() else None,
    }


@router.get(
    "/{run_id}/report",
    summary="Open the Allure report",
    description="Serves the Allure report entry page (index.html) for a run.",
    responses={
        200: {"description": "Allure report HTML"},
        404: {"description": "Report or run not found"},
    },
)
async def get_allure_report_index(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> FileResponse:
    workspace = await _get_run_workspace(run_id, service)
    index_path = _get_allure_report_dir(workspace).joinpath("index.html")
    existing_index = _existing_file_path(index_path)

    if existing_index is None:
        pw_dir = workspace / "artifacts" / "generated-tests" / "playwright"
        test_results_dir = pw_dir / "test-results"
        fallback_tests = _parse_test_results_from_folders(test_results_dir, pw_dir)
        if fallback_tests:
            from app.execution.allure_report_generator import AllureReportGenerator
            generator = AllureReportGenerator()
            generator.generate(
                results_dir=pw_dir / "allure-results",
                output_path=_get_allure_report_dir(workspace),
                project_path=pw_dir if pw_dir.exists() else None,
                fallback_test_results=fallback_tests,
                force_rebuild=True,
            )
            existing_index = _existing_file_path(index_path)

    if existing_index is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Allure report not available for this run",
        )
    return FileResponse(existing_index, media_type="text/html", headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@router.get(
    "/{run_id}/report/download",
    summary="Download Allure report as ZIP archive",
    description="Packages the full interactive HTML Allure report into a downloadable ZIP archive.",
    responses={
        200: {"description": "ZIP file stream"},
        404: {"description": "Report not found"},
    },
)
async def download_allure_report_zip(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> FileResponse:
    workspace = await _get_run_workspace(run_id, service)
    report_dir = _get_allure_report_dir(workspace)

    if not report_dir.exists() or not (report_dir / "index.html").exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Allure report not generated yet for this run",
        )

    zip_path = workspace / "artifacts" / f"allure-report-{run_id}.zip"
    zip_path.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file in report_dir.rglob("*"):
            if file.is_file():
                arcname = file.relative_to(report_dir)
                zf.write(file, arcname=arcname)

    return FileResponse(
        path=zip_path,
        filename=f"allure-report-{run_id}.zip",
        media_type="application/zip",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@router.get(
    "/{run_id}/report/{file_path:path}",
    summary="Serve Allure report asset",
    description="Serves static assets (JS/CSS/data) of the Allure report, validated to stay inside the report directory.",
    responses={
        200: {"description": "Report asset"},
        404: {"description": "Asset or run not found"},
    },
)
async def get_allure_report_asset(
    run_id: UUID,
    file_path: str,
    service: TriggerService = Depends(get_trigger_service),
) -> FileResponse:
    workspace = await _get_run_workspace(run_id, service)
    report_dir = _get_allure_report_dir(workspace)

    target = _resolve_report_file(report_dir, file_path)
    if target is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invalid report asset path",
        )
    existing_target = _existing_file_path(target)
    if existing_target is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report asset not found: {file_path}",
        )
    return FileResponse(existing_target, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})





@router.get(
    "/{run_id}/generated-files",
)
async def get_generated_files(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    project_dir = workspace / "artifacts" / "generated-tests" / "playwright"
    metadata_path = project_dir / "code-generation-metadata.json"
    metadata = _load_json(metadata_path) if metadata_path.exists() else None

    if not project_dir.exists():
        return {"exists": False, "files": [], "metadata": metadata, "project_path": None}

    ignored_names = {".git", ".next", "node_modules", "storage", "artifacts", "__pycache__", ".venv"}

    def build_tree(dir_path: Path, prefix: str = "") -> list[dict]:
        items = []
        try:
            entries = sorted(dir_path.iterdir(), key=lambda p: (not p.is_dir(), p.name))
        except Exception as err:
            logger.warning("build_tree_iterdir_failed path=%s error=%s", dir_path, err)
            return items

        for entry in entries:
            if entry.name.startswith(".") or entry.name in ignored_names:
                continue
            rel = f"{prefix}/{entry.name}" if prefix else entry.name
            try:
                if entry.is_dir():
                    children = build_tree(entry, rel)
                    items.append({"name": entry.name, "type": "directory", "path": rel, "children": children})
                else:
                    size = entry.stat().st_size if entry.exists() else 0
                    items.append({"name": entry.name, "type": "file", "path": rel, "size_bytes": size})
            except Exception as entry_err:
                logger.warning("build_tree_entry_failed path=%s error=%s", entry, entry_err)
                continue
        return items

    files = build_tree(project_dir)
    return {
        "exists": True,
        "files": files,
        "metadata": metadata,
        "project_path": str(project_dir),
    }


@router.get(
    "/{run_id}/generated-files/content",
    summary="Get generated file content",
    description="Returns the content of a generated file from the Playwright project.",
)
async def get_generated_file_content(
    run_id: UUID,
    path: str = Query(..., description="Relative or full file path"),
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    project_dir = workspace / "artifacts" / "generated-tests" / "playwright"

    clean_path = path.replace("\\", "/")
    if "playwright/" in clean_path:
        clean_path = clean_path.split("playwright/")[1]
    
    file_path = project_dir / clean_path
    if not file_path.exists():
        direct_path = Path(path)
        if direct_path.exists() and direct_path.is_file():
            file_path = direct_path

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"File not found: {path}")

    content = file_path.read_text(encoding="utf-8", errors="replace")
    return {
        "path": path,
        "content": content,
        "size_bytes": file_path.stat().st_size,
    }


@router.get(
    "/generated-files/content",
    summary="Get generated file content by path",
    description="Returns content of a generated file when path contains full storage path or relative file path.",
)
async def get_generated_file_content_by_path(
    path: str = Query(..., description="Full or relative file path"),
) -> dict:
    target = Path(path)
    if not target.is_absolute():
        target = Path.cwd() / path

    if not target.exists() or not target.is_file():
        # Search inside storage/runs/*/artifacts/generated-tests/playwright/<path>
        runs_dir = Path.cwd() / "storage" / "runs"
        if runs_dir.exists():
            clean_rel = path.replace("\\", "/")
            if "playwright/" in clean_rel:
                clean_rel = clean_rel.split("playwright/")[1]
            for run_folder in sorted(runs_dir.iterdir(), key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True):
                if not run_folder.is_dir():
                    continue
                candidate = run_folder / "artifacts" / "generated-tests" / "playwright" / clean_rel
                if candidate.exists() and candidate.is_file():
                    target = candidate
                    break

    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"File not found: {path}")

    content = target.read_text(encoding="utf-8", errors="replace")
    return {
        "path": path,
        "content": content,
        "size_bytes": target.stat().st_size,
    }


# ===================================================================
# Generated Playwright Project — ZIP Download
# ===================================================================

# Directory names that must never ship in a generated-project archive. These
# are build/dependency/runtime/execution outputs — not generated source.
# Mirrors the ignore list used by ``GET /generated-files`` plus the Playwright
# execution outputs (large, run-specific) that would bloat a code download.
_GENERATED_CODE_EXCLUDED_DIRS = {
    ".git",
    ".next",
    "node_modules",
    "storage",
    "artifacts",
    "__pycache__",
    ".venv",
    ".pytest_cache",
    ".cache",
    "dist",
    # Playwright execution outputs (not part of the generated source tree)
    "test-results",
    "playwright-report",
    "allure-results",
    "allure-report",
    "traces",
    "videos",
    "screenshots",
}

# Named files that must never ship (platform-internal, not source).
_GENERATED_CODE_EXCLUDED_NAMES = {
    "code-generation-metadata.json",
    "code-generation-ir.json",
    "dependency-graph.json",
}

# Real environment files are always excluded. Only ``.env.example`` (with
# placeholder values) is allowed to ship.
_ENV_NAMES = {".env", ".env.local", ".env.development", ".env.production"}
_ENV_EXAMPLE_NAME = ".env.example"


def _is_real_env_file(name: str) -> bool:
    """True when ``name`` is a real environment file (never shipped).

    ``.env.example`` (and variants) are templates and are allowed.
    """
    if name in _ENV_NAMES:
        return True
    return name.startswith(".env.") and not name.startswith(_ENV_EXAMPLE_NAME)


_SECRET_FILE_SUFFIXES = (".key", ".pem", ".pgp", ".p12", ".pfx", ".token")
_SECRET_FILE_STEM_TOKENS = ("secret", "credential")


def _is_secret_file_name(name: str) -> bool:
    """True when the file ``name`` indicates it carries secrets."""
    lowered = name.lower()
    if lowered.endswith(_SECRET_FILE_SUFFIXES):
        return True
    stem = Path(name).stem.lower()
    return any(token in stem for token in _SECRET_FILE_STEM_TOKENS)


# High-signal secret content patterns (private keys, well-known provider
# tokens). Deliberately strict — avoids flagging placeholder test values.
_SECRET_CONTENT_PATTERNS = (
    re.compile(r"-----BEGIN (?:[A-Z0-9 ]+ )?PRIVATE KEY-----"),
    re.compile(r"AKIA[0-9A-Z]{16}"),                    # AWS access key id
    re.compile(r"\bgh[puorst]_[A-Za-z0-9]{20,}"),       # GitHub tokens
    re.compile(r"xox[baprs]-[A-Za-z0-9-]{10,}"),        # Slack tokens
    re.compile(r"AIza[0-9A-Za-z_\-]{35}"),              # Google API key
    re.compile(r"\bsk-[A-Za-z0-9]{20,}"),               # OpenAI-style API key
)


def _contains_secret(content: bytes) -> bool:
    """True when ``content`` embeds a high-signal secret (private key, token)."""
    text = content.decode("utf-8", errors="ignore")
    return any(pattern.search(text) for pattern in _SECRET_CONTENT_PATTERNS)


def _env_example_placeholder(key: str) -> str:
    """Return a safe placeholder for an environment ``key``."""
    lowered = key.lower()
    if any(
        token in lowered
        for token in ("password", "secret", "token", "key", "credential")
    ):
        return "<your-value>"
    if lowered == "base_url":
        return "http://localhost:3000"
    return "<value>"


def _build_env_example(env_path: Path) -> str | None:
    """Synthesise a placeholder-only ``.env.example`` from a real ``.env``.

    Keeps variable names/comments/documentation but replaces every value with
    a placeholder. Never leaks credentials. Returns ``None`` when missing.
    """
    if not env_path.is_file():
        return None
    try:
        raw = env_path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return None

    lines: list[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            lines.append(line)
            continue
        if "=" not in stripped:
            lines.append(line)
            continue
        key = stripped.split("=", 1)[0].strip()
        lines.append(f"{key}={_env_example_placeholder(key)}")

    if not lines:
        return None
    return "\n".join(lines).rstrip() + "\n"


_GENERATED_CODE_ARCHIVE_ROOT = "generated-tests/playwright"


def _build_generated_code_zip(project_dir: Path) -> bytes | None:
    """Package a generated Playwright project into a sanitised ZIP archive.

    Security guarantees:
    * every entry is resolved and verified to live inside ``project_dir``
      (path traversal / symlink escape is impossible);
    * real ``.env`` files and secret-bearing files are never included;
    * a placeholder ``.env.example`` is always shipped when the project
      references environment variables.

    Returns ZIP bytes, or ``None`` when nothing may be downloaded.
    """
    root_resolved = project_dir.resolve()
    buffer = io.BytesIO()

    env_example_path = project_dir / _ENV_EXAMPLE_NAME
    has_env_example = env_example_path.is_file()
    env_example_content = None if has_env_example else _build_env_example(project_dir / ".env")

    added_files = 0
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in sorted(project_dir.rglob("*")):
            if file_path.is_dir():
                continue

            name = file_path.name
            if name == _ENV_EXAMPLE_NAME:
                continue  # emitted separately below
            if _is_real_env_file(name):
                continue
            if _is_secret_file_name(name):
                continue
            if name in _GENERATED_CODE_EXCLUDED_NAMES:
                continue

            try:
                rel = file_path.relative_to(project_dir)
            except ValueError:
                continue
            # Only evaluate the component sub-tree; ancestors (e.g. the run's
            # workspace under ``storage/runs/*/artifacts``) are never compared.
            if any(part in _GENERATED_CODE_EXCLUDED_DIRS for part in rel.parts):
                continue

            # Path-traversal / symlink-escape guard: the file must resolve to
            # a real location inside the project root.
            try:
                file_path.resolve().relative_to(root_resolved)
            except ValueError:
                logger.warning(
                    "generated_code_download_skipped_outside_path path=%s",
                    file_path,
                )
                continue

            try:
                content = file_path.read_bytes()
            except OSError:
                logger.warning(
                    "generated_code_download_unreadable path=%s",
                    file_path,
                )
                continue

            if content and _contains_secret(content):
                logger.warning(
                    "generated_code_download_skipped_secret_content path=%s",
                    file_path,
                )
                continue

            zf.writestr(
                f"{_GENERATED_CODE_ARCHIVE_ROOT}/{rel.as_posix()}",
                content,
            )
            added_files += 1

        # Always ship a placeholder-only .env.example when the project relies
        # on environment variables. Never the real .env.
        if has_env_example:
            try:
                example_raw = env_example_path.read_bytes()
            except OSError:
                example_raw = b""
            if example_raw and not _contains_secret(example_raw):
                zf.writestr(f"{_GENERATED_CODE_ARCHIVE_ROOT}/{_ENV_EXAMPLE_NAME}", example_raw)
                added_files += 1
        elif env_example_content is not None:
            zf.writestr(f"{_GENERATED_CODE_ARCHIVE_ROOT}/{_ENV_EXAMPLE_NAME}", env_example_content)
            added_files += 1

    if added_files == 0:
        return None

    return buffer.getvalue()


@router.get(
    "/{run_id}/generated-code/download",
    summary="Download generated Playwright project",
    description=(
        "Packages the COMPLETE generated Playwright project for a run into a "
        "downloadable ZIP archive. Secrets are scrubbed — real .env files and "
        "credential files are never included; only a placeholder .env.example "
        "ships with the archive."
    ),
    responses={
        200: {
            "description": "ZIP file stream",
            "content": {"application/zip": {}},
        },
        404: {"description": "Run or generated project not found"},
    },
)
async def download_generated_code(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> StreamingResponse:
    """Download the complete generated Playwright project as a ZIP archive.

    The run ID is the only identifier accepted from the client. Files are read
    exclusively from this run's ``artifacts/generated-tests/playwright``
    directory — arbitrary paths (client-supplied or otherwise) are never used.
    """
    workspace = await _get_run_workspace(run_id, service)
    project_dir = workspace / "artifacts" / "generated-tests" / "playwright"

    if not project_dir.exists() or not project_dir.is_dir():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                f"Generated project not found for run: {run_id}. "
                "Code generation may not have completed."
            ),
        )

    archive_bytes = _build_generated_code_zip(project_dir)
    if archive_bytes is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Generated project is not ready for download for run: {run_id}.",
        )

    filename = f"playwright-generated-code-{run_id}.zip"
    return StreamingResponse(
        io.BytesIO(archive_bytes),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate",
        },
    )


# ===================================================================
# Allure Report — Regeneration
# ===================================================================


@router.post(
    "/{run_id}/report/regenerate",
    summary="Regenerate Allure report",
    description=(
        "Re-runs the Allure report generator for an existing run using the current "
        "allure-results directory.  If the directory is empty the fallback path "
        "synthesises result files from the Playwright JSON output so that the report "
        "contains proper execution details and non-zero durations."
    ),
    responses={
        200: {"description": "Regeneration result"},
        404: {"description": "Run not found"},
    },
)
async def regenerate_allure_report(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    from app.execution.allure_report_generator import AllureReportGenerator

    workspace = await _get_run_workspace(run_id, service)

    exec_artifacts = workspace / "artifacts" / "generated-tests" / "execution-artifacts"
    pw_dir = workspace / "artifacts" / "generated-tests" / "playwright"
    test_results_dir = pw_dir / "test-results"

    allure_results_dir = pw_dir / "allure-results"
    report_dir = _get_allure_report_dir(workspace)

    # Collect fallback test results so the generator can synthesise Allure
    # result files if the allure-playwright reporter did not write any.
    fallback_tests = _parse_test_results_from_folders(test_results_dir, pw_dir)

    # Wipe any previously generated report so the UI always shows fresh data.
    if report_dir.exists():
        shutil.rmtree(report_dir, ignore_errors=True)

    generator = AllureReportGenerator()
    result = generator.generate(
        results_dir=allure_results_dir,
        output_path=report_dir,
        project_path=pw_dir if pw_dir.exists() else None,
        fallback_test_results=fallback_tests or None,
        force_rebuild=True,
    )

    return {
        "run_id": str(run_id),
        "status": result.get("status"),
        "report_available": result.get("status") == "generated",
        "error": result.get("error"),
    }


# ===================================================================
# Allure Report Root Asset Fallback Route
# (MUST remain at the very end of workflow.py to prevent route interception)
# ===================================================================


@router.get(
    "/{run_id}/{asset_dir}/{file_path:path}",
    summary="Serve Allure report root asset",
    description="Serves root-relative Allure SPA assets such as /assets, /data, /widgets, /export, and /history.",
    responses={
        200: {"description": "Report asset"},
        404: {"description": "Asset or run not found"},
    },
)
async def get_allure_root_asset(
    run_id: UUID,
    asset_dir: str,
    file_path: str,
    service: TriggerService = Depends(get_trigger_service),
) -> FileResponse:
    if asset_dir not in {"assets", "data", "widgets", "export", "history"}:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invalid report asset path",
        )

    workspace = await _get_run_workspace(run_id, service)
    report_dir = _get_allure_report_dir(workspace)
    relative_path = f"{asset_dir}/{file_path}"
    target = _resolve_report_file(report_dir, relative_path)
    if target is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invalid report asset path",
        )
    existing_target = _existing_file_path(target)
    if existing_target is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report asset not found: {relative_path}",
        )
    return FileResponse(existing_target, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
