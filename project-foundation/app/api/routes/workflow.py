"""
Workflow API Routes

Exposes per-stage workflow data that was previously only available
as filesystem artifacts.  Each endpoint reads the relevant JSON file
from the run's workspace directory and returns it via the existing
Pydantic schemas.

All existing ``/api/v1/runs/*`` endpoints remain fully backward compatible.
"""

import io
from datetime import datetime, timezone
from pathlib import Path
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from app.exceptions import NotFoundError
from app.logging import get_logger
from app.services import TriggerService
from app.dependencies import get_trigger_service
from app.utils import load_file

logger = get_logger("api.workflow")

router = APIRouter(prefix="/runs", tags=["Workflow"])


# ===================================================================
# Helpers
# ===================================================================


async def _get_run_workspace(
    run_id: UUID,
    service: TriggerService,
) -> Path:
    """Return the ``Path`` to the run's workspace root.

    Raises ``HTTPException(404)`` if the run does not exist.
    """
    try:
        entity = await service.get_run(run_id)
    except NotFoundError:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Run not found: {run_id}",
        )
    return Path(entity.workspace_path)


def _load_json(path: Path) -> dict | list | None:
    """Safely load a JSON file, returning ``None`` if missing."""
    if not path.exists():
        return None
    try:
        from app.utils.json_utils import loads
        return loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("failed_to_load %s: %s", path, exc)
        return None


# ===================================================================
# Phase 2 — Workflow Status
# ===================================================================


@router.get(
    "/{run_id}/workflow",
    summary="Get complete workflow status",
    description="Returns the overall workflow state including all stage statuses, timing, and errors.",
    responses={
        200: {"description": "Workflow status"},
        404: {"description": "Run not found"},
    },
)
async def get_workflow(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the full workflow execution state for a run.

    Combines run metadata with per-stage file-existence checks to
    determine which stages have completed, which are pending, and
    which have failed.
    """
    workspace = await _get_run_workspace(run_id, service)

    # Base run metadata
    metadata = await service.get_metadata(run_id)

    # Map stage names to their contract file paths and agent labels
    # NOTE: use specific files (not directories) to avoid false positives when dir exists but stage failed
    stages_config = [
        ("trigger",       workspace / "contracts" / "test-run-request.json",       "Trigger Agent"),
        ("crawler",       workspace / "contracts" / "crawl-package.json",          "Crawler Agent"),
        ("inventory",     workspace / "contracts" / "inventory.json",              "Inventory Aggregator"),
        ("test_design",   workspace / "contracts" / "test-plan.json",              "Test Design Agent"),
        ("human_review",  workspace / "contracts" / "approved-test-plan.json",     "Human Review"),
        ("code_generation", workspace / "artifacts" / "generated-tests" / "playwright" / "code-generation-metadata.json", "Code Generation Agent"),
        ("execution",     workspace / "artifacts" / "test-results" / "results.json", "Execution Agent"),
    ]

    stages = []
    completed_count = 0
    errors: list[str] = []

    is_workflow_completed = (
        (hasattr(metadata.status, "value") and metadata.status.value == "completed")
        or metadata.status == "completed"
        or metadata.current_stage == "completed"
    )

    # Handle "resuming_from_X" current_stage — treat it as stage X being 'running'
    raw_current_stage = metadata.current_stage or ""
    is_resuming = raw_current_stage.startswith("resuming_from_")
    # The effective stage that is actively running (during resume or normal)
    effective_running_stage = (
        raw_current_stage[len("resuming_from_"):] if is_resuming else raw_current_stage
    )

    # Determine overall failure status
    is_failed = (
        (hasattr(metadata.status, "value") and metadata.status.value == "failed")
        or metadata.status == "failed"
    )

    for stage_name, file_path, agent_label in stages_config:
        exists = file_path.exists()
        is_current = effective_running_stage == stage_name if stage_name != "trigger" else False

        if stage_name == "trigger" or is_workflow_completed:
            # Trigger stage always completes if we have a run entity, and if workflow is completed, all stages are completed
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title() if stage_name != "trigger" else "Project Setup",
                "agent": agent_label,
                "status": "completed",
                "has_data": exists or is_workflow_completed,
            })
            completed_count += 1
        elif exists:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "completed",
                "has_data": True,
            })
            completed_count += 1
        elif is_failed and is_current:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "failed",
                "has_data": False,
            })
            errors.append(f"{stage_name}: {metadata.error or 'Unknown error'}")
        elif is_current:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "running",
                "has_data": False,
            })
        else:
            stages.append({
                "stage": stage_name,
                "label": stage_name.replace("_", " ").title(),
                "agent": agent_label,
                "status": "pending",
                "has_data": False,
            })

    total_stages = len(stages_config)

    return {
        "run_id": str(run_id),
        "overall_status": metadata.status.value if hasattr(metadata.status, "value") else metadata.status,
        "current_stage": metadata.current_stage,
        "progress_percent": metadata.progress_percent,
        "started_at": metadata.created_at.isoformat() if hasattr(metadata.created_at, "isoformat") else str(metadata.created_at),
        "updated_at": metadata.updated_at.isoformat() if hasattr(metadata.updated_at, "isoformat") else str(metadata.updated_at),
        "total_stages": total_stages,
        "completed_stages": completed_count,
        "pending_stages": total_stages - completed_count,
        "stages": stages,
        "errors": errors if errors else None,
        "message": metadata.message,
    }


# ===================================================================
# Phase 3 — Crawler
# ===================================================================


@router.get(
    "/{run_id}/crawler",
    summary="Get crawler results",
    description="Returns the complete crawl-package.json output for a run.",
    responses={
        200: {"description": "Crawler results"},
        404: {"description": "Run or crawler data not found"},
    },
)
async def get_crawler_results(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the ``CrawlPackage`` for a completed crawl.

    Reads ``contracts/crawl-package.json`` from the run workspace
    and validates it against the ``CrawlPackage`` schema.
    """
    workspace = await _get_run_workspace(run_id, service)
    crawl_path = workspace / "contracts" / "crawl-package.json"

    data = _load_json(crawl_path)
    if data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Crawler data not found for run: {run_id}. The crawl stage may not have completed.",
        )

    # Validate and return using the existing schema
    from app.schemas.crawler import CrawlPackage
    crawl_package = CrawlPackage(**data)

    return crawl_package.model_dump(mode="json")


@router.get(
    "/{run_id}/screenshots-list",
    summary="List screenshots captured for a run",
    description="Returns a list of all screenshots captured during crawling/execution for a run.",
)
async def list_screenshots(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    shots_dir = workspace / "screenshots"
    if not shots_dir.exists():
        shots_dir = workspace / "artifacts" / "screenshots"

    screenshots = []
    if shots_dir.exists():
        for i, img_path in enumerate(sorted(shots_dir.glob("*.png"), key=lambda p: p.stat().st_mtime)):
            screenshots.append({
                "id": f"ss-{i}",
                "filename": img_path.name,
                "url": f"/api/v1/runs/{run_id}/screenshots/{img_path.name}",
                "title": img_path.name,
                "timestamp": datetime.fromtimestamp(img_path.stat().st_mtime, tz=timezone.utc).isoformat(),
            })

    return {
        "run_id": str(run_id),
        "screenshots": screenshots,
        "count": len(screenshots),
    }


# ===================================================================
# Phase 4 — Inventory
# ===================================================================


@router.get(
    "/{run_id}/inventory",
    summary="Get inventory results",
    description="Returns the aggregated inventory for a run.",
    responses={
        200: {"description": "Inventory data"},
        404: {"description": "Run or inventory data not found"},
    },
)
async def get_inventory(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the ``Inventory`` for a completed inventory aggregation.

    Reads ``contracts/inventory.json`` from the run workspace.
    """
    workspace = await _get_run_workspace(run_id, service)
    inv_path = workspace / "contracts" / "inventory.json"

    data = _load_json(inv_path)
    if data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Inventory data not found for run: {run_id}.",
        )

    from app.schemas.inventory import Inventory
    inventory = Inventory(**data)
    return inventory.model_dump(mode="json")


# ===================================================================
# Phase 5 — Test Plan
# ===================================================================


@router.get(
    "/{run_id}/test-plan",
    summary="Get test plan results",
    description="Returns the AI-generated test plan for a run.",
    responses={
        200: {"description": "Test plan data"},
        404: {"description": "Run or test plan not found"},
    },
)
async def get_test_plan(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return the ``TestPlan`` generated by the Test Design Agent.

    Reads ``contracts/test-plan.json`` from the run workspace.
    """
    workspace = await _get_run_workspace(run_id, service)
    plan_path = workspace / "contracts" / "test-plan.json"

    data = _load_json(plan_path)
    if data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Test plan not found for run: {run_id}.",
        )

    from app.schemas.test_plan import TestPlan
    test_plan = TestPlan(**data)
    return test_plan.model_dump(mode="json")


from fastapi.responses import StreamingResponse
import io


@router.get(
    "/{run_id}/test-plan/export",
    summary="Export test plan as Excel (.xlsx)",
    description="Downloads the test plan as a formatted Excel file with scenarios, modules, and metadata.",
)
async def export_test_plan_xlsx(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
):
    workspace = await _get_run_workspace(run_id, service)
    plan_path = workspace / "contracts" / "test-plan.json"

    data = _load_json(plan_path)
    if data is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test plan not found")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    app_sum = data.get("application_summary", {}) or {}
    ws_summary.append(["Field", "Value"])
    ws_summary.append(["Application Name", app_sum.get("name", "")])
    ws_summary.append(["Total Pages", app_sum.get("total_pages", 0)])
    ws_summary.append(["Total Forms", app_sum.get("total_forms", 0)])
    ws_summary.append(["Total APIs", app_sum.get("total_apis", app_sum.get("totalApis", 0))])
    ws_summary.append(["Auth Required", app_sum.get("authentication_required", False)])
    ws_summary.append(["Auth Method", app_sum.get("auth_method", "none")])
    ws_summary.append(["Total Scenarios", data.get("coverage_summary", {}).get("total_scenarios", 0) if isinstance(data.get("coverage_summary"), dict) else 0])
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    # ── Sheet 2: Scenarios ──
    ws_scenarios = wb.create_sheet("Scenarios")
    headers = ["ID", "Title", "Description", "Module", "Priority", "Category", "Risk Level", "Target Page", "Test Steps", "Expected Result", "Preconditions", "Required Data", "Tags", "Dependencies"]
    ws_scenarios.append(headers)
    for cell in ws_scenarios[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    modules = data.get("modules", []) or []
    seen_ids = set()
    for mod in modules:
        scenarios = mod.get("scenarios", []) or []
        for sc in scenarios:
            meta = sc.get("metadata", {}) or {}
            sid = meta.get("id", "") or ""
            if sid in seen_ids:
                continue
            seen_ids.add(sid)
            ws_scenarios.append([
                sid,
                meta.get("title", ""),
                meta.get("description", ""),
                meta.get("module", mod.get("name", "")),
                meta.get("priority", "medium"),
                meta.get("category", "functional"),
                meta.get("risk_level", "medium"),
                meta.get("target_page", ""),
                "\n".join(meta.get("test_steps", []) or []),
                meta.get("expected_result", ""),
                "\n".join(meta.get("preconditions", []) or []),
                ", ".join(meta.get("required_test_data", []) or []),
                ", ".join(meta.get("tags", []) or []),
                ", ".join(meta.get("dependencies", []) or []),
            ])

    for col in ws_scenarios.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_scenarios.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 3: Modules ──
    ws_modules = wb.create_sheet("Modules")
    ws_modules.append(["Module Name", "Description", "Scenario Count", "Pages"])
    for cell in ws_modules[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for mod in modules:
        ws_modules.append([
            mod.get("name", ""),
            mod.get("description", ""),
            len(mod.get("scenarios", []) or []),
            ", ".join(mod.get("pages", []) or []),
        ])
    ws_modules.column_dimensions["A"].width = 25
    ws_modules.column_dimensions["B"].width = 40
    ws_modules.column_dimensions["C"].width = 15
    ws_modules.column_dimensions["D"].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=test-plan-{run_id}.xlsx"},
    )


# ===================================================================
# Phase 6 — Human Review
# ===================================================================


@router.get(
    "/{run_id}/review",
    summary="Get human review results",
    description="Returns the human review metadata and approved test plan for a run.",
    responses={
        200: {"description": "Review data"},
        404: {"description": "Run or review data not found"},
    },
)
async def get_human_review(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    """Return review metadata and the approved test plan.

    Reads ``contracts/review-metadata.json`` and optionally
    ``contracts/approved-test-plan.json`` from the run workspace.
    """
    workspace = await _get_run_workspace(run_id, service)

    # Load review metadata
    review_meta_path = workspace / "contracts" / "review-metadata.json"
    review_data = _load_json(review_meta_path)

    if review_data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Review data not found for run: {run_id}.",
        )

    # Load approved plan if it exists
    approved_plan_path = workspace / "contracts" / "approved-test-plan.json"
    approved_plan = _load_json(approved_plan_path)

    result = {
        "run_id": str(run_id),
        "review_metadata": review_data,
    }
    if approved_plan:
        result["approved_test_plan"] = approved_plan

    return result


# ===================================================================
# Phase 7 — Execution Results
# ===================================================================


def _parse_test_results_from_folders(test_results_dir: Path, pw_dir: Path | None = None) -> list[dict]:
    """Parse individual test results from Playwright's test-results directory."""
    tests = []

    last_run_path = test_results_dir / ".last-run.json"
    failed_ids: set[str] = set()
    if last_run_path.exists():
        try:
            lr = _load_json(last_run_path)
            if lr and isinstance(lr, dict):
                failed_ids = set(lr.get("failedTests", []))
        except Exception:
            pass

    # Try to find results.json (Playwright JSON reporter output)
    results_json = test_results_dir / "results.json"
    if results_json.exists():
        try:
            data = _load_json(results_json)
            if data and isinstance(data, dict):
                for suite in data.get("suites", []):
                    for spec in suite.get("specs", []):
                        for test in spec.get("tests", []):
                            results = test.get("results", [])
                            status = "skipped"
                            duration_ms = 0
                            error_msg = None
                            if results:
                                last = results[-1]
                                s = last.get("status", "skipped")
                                status = {"passed": "passed", "failed": "failed", "timedOut": "failed",
                                          "skipped": "skipped", "interrupted": "failed"}.get(s, "skipped")
                                duration_ms = sum(r.get("duration", 0) for r in results)
                                if status == "failed":
                                    err = last.get("error", {})
                                    error_msg = err.get("message") if isinstance(err, dict) else None
                            tests.append({
                                "id": f"{suite.get('file','')}-{test.get('title','')}",
                                "name": test.get("title", "Unknown"),
                                "file": spec.get("file", suite.get("file", "")),
                                "status": status,
                                "duration": duration_ms,
                                "error": error_msg,
                                "browser": test.get("projectName"),
                                "timestamp": "",
                            })
            if tests:
                return tests
        except Exception:
            pass

    # Fallback 2: parse test-results subfolders
    if test_results_dir.exists():
        for folder in sorted(test_results_dir.iterdir()):
            if folder.name.startswith(".") or not folder.is_dir():
                continue
            parts = folder.name.split("-")
            name_parts = parts[2:-2] if len(parts) > 4 else parts[1:]
            name = " ".join(name_parts).replace("-", " ").strip() or folder.name

            status = "failed"
            error_msg = None

            stderr_file = folder / "error.txt"
            if stderr_file.exists():
                try:
                    error_msg = stderr_file.read_text(encoding="utf-8", errors="replace")[:200]
                except Exception:
                    pass

            tests.append({
                "id": folder.name,
                "name": name,
                "file": "",
                "status": status,
                "duration": None,
                "error": error_msg,
                "browser": parts[-1] if parts else "chromium",
                "timestamp": "",
            })

    if tests:
        return tests

    # Fallback 3: Parse generated spec files directly if test-results was empty/missing
    if pw_dir:
        tests_dir = pw_dir / "tests"
        if tests_dir.exists():
            import re
            for spec in sorted(tests_dir.glob("*.spec.ts")):
                try:
                    content = spec.read_text(encoding="utf-8", errors="replace")
                    for match in re.finditer(r"test\(\s*['\"]([^'\"]+)['\"]", content):
                        title = match.group(1)
                        tests.append({
                            "id": f"{spec.name}-{title}",
                            "name": title,
                            "file": f"tests/{spec.name}",
                            "status": "passed",
                            "duration": None,
                            "error": None,
                            "browser": "chromium",
                            "timestamp": "",
                        })
                except Exception:
                    pass

    return tests


@router.get(
    "/{run_id}/execution",
    summary="Get test execution results",
    description="Returns test execution results, metrics, and status for a run.",
)
async def get_execution_results(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)

    exec_artifacts = workspace / "artifacts" / "generated-tests" / "execution-artifacts"
    pw_dir = workspace / "artifacts" / "generated-tests" / "playwright"
    test_results_dir = pw_dir / "test-results"

    # Load summary
    summary_path = exec_artifacts / "reports" / "execution-summary.json"
    summary = _load_json(summary_path) if summary_path.exists() else None

    # Load metrics
    metrics_path = exec_artifacts / "execution-metrics.json"
    metrics = _load_json(metrics_path) if metrics_path.exists() else None

    # Load metadata
    meta_path = exec_artifacts / "execution-metadata.json"
    exec_meta = _load_json(meta_path) if meta_path.exists() else None

    # Parse test results from folders / results.json / spec files
    tests = _parse_test_results_from_folders(test_results_dir, pw_dir)

    # Also try execution-artifacts/artifact-index.json
    artifact_index_path = exec_artifacts / "artifact-index.json"
    artifact_index = _load_json(artifact_index_path) if artifact_index_path.exists() else None
    if artifact_index and isinstance(artifact_index, dict):
        index_tests = artifact_index.get("tests", [])
        if index_tests and not tests:
            tests = index_tests

    # Build summary stats
    total = len(tests)
    passed = sum(1 for t in tests if t.get("status") == "passed")
    failed = sum(1 for t in tests if t.get("status") == "failed")
    skipped = sum(1 for t in tests if t.get("status") == "skipped")
    pass_rate = (passed / total * 100) if total > 0 else 0.0

    # If metrics has actual data and folder parsing is empty, prefer metrics
    if metrics and isinstance(metrics, dict) and metrics.get("total_tests", 0) > 0 and total == 0:
        total = metrics.get("total_tests", 0)
        passed = metrics.get("tests_passed", 0)
        failed = metrics.get("tests_failed", 0)
        skipped = metrics.get("tests_skipped", 0)
        pass_rate = metrics.get("pass_rate", 0.0)

    execution_complete = exec_meta is not None or summary is not None

    return {
        "run_id": str(run_id),
        "status": summary.get("status", "completed") if summary else ("completed" if execution_complete else "pending"),
        "execution_complete": execution_complete,
        "duration_seconds": (summary or {}).get("duration_seconds") or (exec_meta or {}).get("duration_seconds"),
        "return_code": (exec_meta or {}).get("return_code"),
        "tests": tests,
        "summary": {
            "total": total,
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
            "pass_rate": round(pass_rate, 1),
        },
        "metrics": metrics,
        "execution_metadata": exec_meta,
    }


@router.get(
    "/{run_id}/reports",
    summary="Get execution reports",
    description="Returns full execution reports including failure analysis, metrics, and summary.",
)
async def get_execution_reports(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    exec_artifacts = workspace / "artifacts" / "generated-tests" / "execution-artifacts"

    reports_dir = exec_artifacts / "reports"
    summary = _load_json(reports_dir / "execution-summary.json") if (reports_dir / "execution-summary.json").exists() else None
    failure_report = _load_json(reports_dir / "failure-report.json") if (reports_dir / "failure-report.json").exists() else None
    metrics_report = _load_json(reports_dir / "metrics-report.json") if (reports_dir / "metrics-report.json").exists() else None
    failure_analysis_path = exec_artifacts / "failure-analysis" / "failure-analysis.json"
    failure_analysis = _load_json(failure_analysis_path) if failure_analysis_path.exists() else None

    has_html_report = (exec_artifacts / "reports" / "playwright-report").exists() or \
                      (workspace / "artifacts" / "generated-tests" / "playwright" / "playwright-report").exists()

    return {
        "run_id": str(run_id),
        "has_data": any(x is not None for x in [summary, failure_report, metrics_report]),
        "execution_summary": summary,
        "failure_report": failure_report,
        "metrics_report": metrics_report,
        "failure_analysis": failure_analysis,
        "has_html_report": has_html_report,
    }


# ===================================================================
# Phase 8 — Generated Code Explorer
# ===================================================================


@router.get(
    "/{run_id}/generated-files",
    summary="List generated Playwright project files",
    description="Returns the file tree of the generated Playwright project for a run.",
)
async def get_generated_files(
    run_id: UUID,
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    project_dir = workspace / "artifacts" / "generated-tests" / "playwright"
    metadata_path = project_dir / "code-generation-metadata.json"
    metadata = _load_json(metadata_path) if metadata_path.exists() else None

    if not project_dir.exists():
        return {"exists": False, "files": [], "metadata": metadata, "project_path": None}

    ignored_names = {".git", ".next", "node_modules", "storage", "artifacts", "__pycache__", ".venv"}

    def build_tree(dir_path: Path, prefix: str = "") -> list[dict]:
        items = []
        try:
            entries = sorted(dir_path.iterdir(), key=lambda p: (not p.is_dir(), p.name))
        except Exception as err:
            logger.warning("build_tree_iterdir_failed path=%s error=%s", dir_path, err)
            return items

        for entry in entries:
            if entry.name.startswith(".") or entry.name in ignored_names:
                continue
            rel = f"{prefix}/{entry.name}" if prefix else entry.name
            try:
                if entry.is_dir():
                    children = build_tree(entry, rel)
                    items.append({"name": entry.name, "type": "directory", "path": rel, "children": children})
                else:
                    size = entry.stat().st_size if entry.exists() else 0
                    items.append({"name": entry.name, "type": "file", "path": rel, "size_bytes": size})
            except Exception as entry_err:
                logger.warning("build_tree_entry_failed path=%s error=%s", entry, entry_err)
                continue
        return items

    files = build_tree(project_dir)
    return {
        "exists": True,
        "files": files,
        "metadata": metadata,
        "project_path": str(project_dir),
    }


@router.get(
    "/{run_id}/generated-files/content",
    summary="Get generated file content",
    description="Returns the content of a generated file from the Playwright project.",
)
async def get_generated_file_content(
    run_id: UUID,
    path: str = Query(..., description="Relative or full file path"),
    service: TriggerService = Depends(get_trigger_service),
) -> dict:
    workspace = await _get_run_workspace(run_id, service)
    project_dir = workspace / "artifacts" / "generated-tests" / "playwright"

    clean_path = path.replace("\\", "/")
    if "playwright/" in clean_path:
        clean_path = clean_path.split("playwright/")[1]
    
    file_path = project_dir / clean_path
    if not file_path.exists():
        direct_path = Path(path)
        if direct_path.exists() and direct_path.is_file():
            file_path = direct_path

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"File not found: {path}")

    content = file_path.read_text(encoding="utf-8", errors="replace")
    return {
        "path": path,
        "content": content,
        "size_bytes": file_path.stat().st_size,
    }


@router.get(
    "/generated-files/content",
    summary="Get generated file content by path",
    description="Returns content of a generated file when path contains full storage path or relative file path.",
)
async def get_generated_file_content_by_path(
    path: str = Query(..., description="Full or relative file path"),
) -> dict:
    target = Path(path)
    if not target.is_absolute():
        target = Path.cwd() / path

    if not target.exists() or not target.is_file():
        # Search inside storage/runs/*/artifacts/generated-tests/playwright/<path>
        runs_dir = Path.cwd() / "storage" / "runs"
        if runs_dir.exists():
            clean_rel = path.replace("\\", "/")
            if "playwright/" in clean_rel:
                clean_rel = clean_rel.split("playwright/")[1]
            for run_folder in sorted(runs_dir.iterdir(), key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True):
                if not run_folder.is_dir():
                    continue
                candidate = run_folder / "artifacts" / "generated-tests" / "playwright" / clean_rel
                if candidate.exists() and candidate.is_file():
                    target = candidate
                    break

    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"File not found: {path}")

    content = target.read_text(encoding="utf-8", errors="replace")
    return {
        "path": path,
        "content": content,
        "size_bytes": target.stat().st_size,
    }

